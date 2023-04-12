from openpyxl import load_workbook
import parsingxml
import winsound

wb = load_workbook('test.xlsx')
ws = wb.active

# 3부터 13015번까지
for i in range(3, 13016):
    number = str(i)
    result = parsingxml.loadDataFromUrl(ws["D" + number].value)
    print(number)
    ws["E" + number].value = result[0]
    ws["F" + number].value = result[1]
    ws["G" + number].value = result[2]
    ws["H" + number].value = result[3]

wb.save('resultdata2.xlsx')

winsound.PlaySound("SystemHand", winsound.SND_ALIAS)